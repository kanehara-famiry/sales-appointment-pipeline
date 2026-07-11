#!/opt/data/.venv_sales/bin/python3
"""
パイプライン結果書き込みスクリプト。
Agent ⑧ (Proposal Writer) の出力を受け取り、スプレッドシートを更新する。

Usage: pipeline_writer.py '<json_data>'
"""
import sys, json, os
sys.path.insert(0, "/opt/data/.venv_sales/lib/python3.13/site-packages")
import openpyxl
from datetime import datetime

SHEET_PATH = "/opt/data/sales_appointment_prep_package.xlsx"

def write_sheet(data):
    """Update spreadsheet with pipeline results."""
    wb = openpyxl.load_workbook(SHEET_PATH)
    ws = wb["入力管理"]
    
    apo_id = data.get("apo_id")
    sheet_updates = data.get("sheet_updates", [])
    new_status = data.get("status", "レビュー待ち")
    confirm_count = data.get("confirm_count", 0)
    
    # Find row
    found_row = None
    for row_idx in range(2, ws.max_row + 1):
        cell_val = ws.cell(row=row_idx, column=1).value
        if cell_val and str(cell_val).strip() == apo_id:
            found_row = row_idx
            break
    
    if not found_row:
        print(json.dumps({"success": False, "error": f"Row for {apo_id} not found"}))
        return False
    
    # Column mapping
    col_map = {
        "M": 13, "N": 14, "O": 15, "P": 16, "Q": 17, "R": 18,
        "S": 19, "T": 20, "U": 21, "V": 22, "W": 23, "X": 24, "Y": 25,
        "Z": 26, "AA": 27, "AB": 28, "AC": 29, "AD": 30, "AE": 31, "AF": 32,
        "AG": 33, "AH": 34, "AI": 35, "AJ": 36
    }
    
    # Apply updates
    for update in sheet_updates:
        col = update.get("col")
        value = update.get("value")
        if col and value is not None and col in col_map:
            ws.cell(row=found_row, column=col_map[col], value=value)
    
    # Status (H=8)
    ws.cell(row=found_row, column=8, value=new_status)
    
    # Confirm count (K=11)
    ws.cell(row=found_row, column=11, value=confirm_count)
    
    # Timestamp (AJ=36)
    ws.cell(row=found_row, column=36, value=datetime.now().strftime("%Y/%m/%d %H:%M"))
    
    wb.save(SHEET_PATH)
    wb.close()
    print(json.dumps({"success": True, "row": found_row, "status": new_status}))
    return True

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(json.dumps({"success": False, "error": "No data provided"}))
        sys.exit(1)
    data = json.loads(sys.argv[1])
    write_sheet(data)
