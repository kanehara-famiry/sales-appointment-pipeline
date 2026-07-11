#!/opt/data/.venv_sales/bin/python3
"""
チェックスクリプト: スプレッドシートをチェックし、新規アポ（未着手）があればJSONで出力する。
なければ空のJSON配列を出力する（cronが静かに終了）。
"""
import sys, json
sys.path.insert(0, "/opt/data/.venv_sales/lib/python3.13/site-packages")
import openpyxl

SHEET_PATH = "/opt/data/sales_appointment_prep_package.xlsx"
wb = openpyxl.load_workbook(SHEET_PATH, data_only=True)
ws = wb["入力管理"]

pending = []
for row_idx in range(2, ws.max_row + 1):
    apo_id = ws.cell(row=row_idx, column=1).value
    status = ws.cell(row=row_idx, column=8).value  # H列 = 準備ステータス
    if apo_id and (status is None or str(status).strip() == "" or str(status).strip() == "未着手"):
        row_data = {}
        headers = ["A","B","C","D","E","F","G","H","I","J","K","L","M","N","O","P","Q","R","S","T","U","V","W","X","Y","Z","AA","AB","AC","AD","AE","AF","AG","AH","AI","AJ"]
        for col_idx, col_letter in enumerate(headers, 1):
            v = ws.cell(row=row_idx, column=col_idx).value
            if v is not None:
                from datetime import datetime as dt
                if isinstance(v, dt):
                    v = v.strftime("%Y/%m/%d %H:%M")
                row_data[col_letter] = str(v).strip() if not isinstance(v, (int, float)) else v
        row_data["_row"] = row_idx  # 行番号を記録
        pending.append(row_data)

if pending:
    print(json.dumps(pending, ensure_ascii=False, default=str))
else:
    print("[]")  # 空の配列 = 何もしない
