#!/opt/data/.venv_sales/bin/python3
"""
Mark an appointment's status in the spreadsheet.
Used by the pipeline orchestrator to update progress.

Usage: status_marker.py <apo_id> <status>
"""
import sys, json
sys.path.insert(0, "/opt/data/.venv_sales/lib/python3.13/site-packages")
import openpyxl
from datetime import datetime

SHEET_PATH = "/opt/data/sales_appointment_prep_package.xlsx"

apo_id = sys.argv[1]
new_status = sys.argv[2]

wb = openpyxl.load_workbook(SHEET_PATH)
ws = wb["入力管理"]

found = False
for row_idx in range(2, ws.max_row + 1):
    cell_val = ws.cell(row=row_idx, column=1).value
    if cell_val and str(cell_val).strip() == apo_id:
        ws.cell(row=row_idx, column=8, value=new_status)  # H=status
        ws.cell(row=row_idx, column=36, value=datetime.now().strftime("%Y/%m/%d %H:%M"))  # AJ=timestamp
        found = True
        break

if found:
    wb.save(SHEET_PATH)
    wb.close()
    print(json.dumps({"success": True, "apo_id": apo_id, "status": new_status}))
else:
    wb.close()
    print(json.dumps({"success": False, "error": f"Appointment {apo_id} not found"}))
    sys.exit(1)
