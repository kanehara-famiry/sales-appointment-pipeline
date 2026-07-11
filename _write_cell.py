#!/opt/data/.venv_sales/bin/python3
"""Write a cell value to sheet1 (入力管理). Takes JSON from argv[1]: {"apo_id": "...", "col": "H", "value": "..."}
Finds the row by matching column A (アポID)."""
import sys, json
sys.path.insert(0, "/opt/data/.venv_sales/lib/python3.13/site-packages")
import openpyxl
from copy import copy

SHEET_PATH = "/opt/data/sales_appointment_prep_package.xlsx"

data = json.loads(sys.argv[1])
apo_id = data["apo_id"]
col = data["col"]
value = data["value"]

wb = openpyxl.load_workbook(SHEET_PATH)
ws = wb["入力管理"]

# Find the row with matching apo_id
found = False
for row_idx in range(2, ws.max_row + 1):
    cell_val = ws.cell(row=row_idx, column=1).value
    if cell_val and str(cell_val).strip() == apo_id:
        col_idx = openpyxl.utils.column_index_from_string(col)
        ws.cell(row=row_idx, column=col_idx, value=value)
        found = True
        break

if not found:
    print(json.dumps({"success": False, "error": f"Appointment {apo_id} not found"}))
    sys.exit(1)

wb.save(SHEET_PATH)
wb.close()
print(json.dumps({"success": True}))
