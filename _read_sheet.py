#!/opt/data/.venv_sales/bin/python3
"""Read sheet1 (入力管理) and output JSON to stdout."""
import sys, json
sys.path.insert(0, "/opt/data/.venv_sales/lib/python3.13/site-packages")
import openpyxl
from datetime import datetime

SHEET_PATH = "/opt/data/sales_appointment_prep_package.xlsx"
wb = openpyxl.load_workbook(SHEET_PATH, data_only=True)
ws = wb["入力管理"]

# Collect all rows
headers = []
rows = []
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False)):
    vals = {}
    for cell in row:
        col_letter = cell.column_letter
        v = cell.value
        if v is not None:
            # Convert datetime objects to string
            if isinstance(v, datetime):
                v = v.strftime("%Y/%m/%d %H:%M")
            vals[col_letter] = str(v).strip() if not isinstance(v, (int, float)) else v
    if i == 0:
        headers = vals
    else:
        if vals.get("A"):  # has an ID
            rows.append(vals)

print(json.dumps(rows, ensure_ascii=False, default=str))
