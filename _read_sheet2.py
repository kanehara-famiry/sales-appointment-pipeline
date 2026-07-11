#!/opt/data/.venv_sales/bin/python3
"""Read sheet2 (エージェント分担表) and output JSON to stdout."""
import sys, json
sys.path.insert(0, "/opt/data/.venv_sales/lib/python3.13/site-packages")
import openpyxl

SHEET_PATH = "/opt/data/sales_appointment_prep_package.xlsx"
wb = openpyxl.load_workbook(SHEET_PATH, data_only=True)
ws = wb["エージェント分担表"]

agents = []
for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False)):
    vals = {}
    for cell in row:
        col = cell.column_letter
        v = cell.value
        if v is not None:
            vals[col] = str(v).strip()
    if vals.get("A"):
        agents.append({
            "agent": vals.get("A", ""),
            "role": vals.get("B", ""),
            "input": vals.get("C", ""),
            "output": vals.get("D", ""),
            "completion": vals.get("E", "")
        })

print(json.dumps(agents, ensure_ascii=False))
